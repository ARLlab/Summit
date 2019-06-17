import pandas as pd
from openpyxl import load_workbook

from datetime import datetime
from pathlib import Path
import datetime as dt
from summit_methane import GcRun
from summit_core import connect_to_db
from summit_core import methane_dir

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename] into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Adapted From:
        https://stackoverflow.com/questions/38074678/append-existing-excel-sheet-with-new-dataframe-using-python-pandas
    :param filename: File path or existing ExcelWriter (Example: '/path/to/file.xlsx')
    :param df: dataframe to save to workbook
    :param sheet_name: Name of sheet which will contain DataFrame. (default: 'Sheet1')
    :param startrow: upper left cell row to dump data frame. Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
    :param truncate_sheet: truncate (remove and recreate) [sheet_name] before writing DataFrame to Excel file
    :param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()` [can be dictionary]
    :return: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0
    else:
        # If there's a start row, don't write the header when appending
        if not to_excel_kwargs:
            to_excel_kwargs = {'header': False}
        else:
            to_excel_kwargs['header'] = False

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def set_formula_in_row(ws, num, row, mr_col=1):

    assert mr_col in [1,2], "Invalid mixing ratio column. It must either 1 or 2"

    std_relnum = 1 if mr_col is 1 else 3
    # if it's the first mixing ratio column, the standard will be in the second row (0-indexed: 1)
    # if it's the second mixing ratio columnm, the standard will be in the fourth row (0-indexed: 3)

    """
    col1  | col2
    -------------
    samp1 | samp2
    std1  | samp4
    samp5 | samp6
    samp7 | std2
    samp9 | samp10
    """
    standard_div_line = 2 if mr_col is 1 else 1
    # samples 1-5 (excluding the standardd) are quantified using the first standard (sample 3)
    # samples 6-10 (excluding the stnadrad) are quantified using the second standard (sample 8)
    # so, in column 1, every sample up to (0-indexed) 2 should be quantified with standard 1, and
    #  everything after with standard 2. In column 2, that number changes to 1

    relnum = num % 5
    # num is 0-indexed, relnum is the position in this group of 5 rows (one run is 5r x 2c for 10 total runs)
    if relnum is std_relnum: return ws  # skip the standard for this column

    for cell in row:
        if cell.value is None:  # assume cells with some value have been modified and should not be changed
            rownum = cell.row  # retrieve the real row number
            pa_cell = f'C{rownum}' if mr_col is 1 else f'D{rownum}'
            # the peak area for a mixing ratio cell will always be C for column 1 and D for column 2, always same row

            if relnum <= standard_div_line:  # this is should be quantified by standard 1, in this column
                std_pa_cell = f'C{rownum - relnum + 1}'
            else:  # it should be quantified by standard 2, in the next column
                std_pa_cell = f'D{rownum - relnum + 3}'

            cell.value = f'={pa_cell}/{std_pa_cell} * 2067.16'

            if relnum is 0 and mr_col is 1:  # the first line of every 5-row batch needs additional statistics added
                # this does not need to be done twice

                run_range = f'E{rownum}:F{rownum+4}'  # all mixing cells in the run
                std_range = f'C{rownum+1}, D{rownum+3}'  # the two standards

                run_median_cell = ws[f'G{rownum}']
                run_rsd_cell = ws[f'H{rownum}']
                std_med_cell = ws[f'I{rownum}']
                std_rsd_cell = ws[f'J{rownum}']

                run_rsd_cell.number_format = '0.00%'
                std_rsd_cell.number_format = '0.00%'

                run_median_cell.value = f'=MEDIAN({run_range})'
                run_rsd_cell.value = f'=STDEV({run_range})/{run_median_cell.coordinate}'
                std_med_cell.value = f'=MEDIAN({std_range})'
                std_rsd_cell.value = f'=STDEV({std_range})/{std_med_cell.coordinate}'
    return ws


engine, session = connect_to_db('sqlite:///summit_methane.sqlite', methane_dir)

three_days_ago = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) - dt.timedelta(days=6)
runs_for_three_days = session.query(GcRun).filter(GcRun.date > three_days_ago).all()

col_list = ['date', 'filename', 'peak1', 'peak2', 'mr1', 'mr2', 'run_median', 'run_rsd', 'std_median', 'std_rsd']

file = Path(r'C:\Users\arl\Desktop\Summit Processing\Summit\processors\summit_methane_processor\tests\test.xlsx')

master_df = pd.DataFrame(index=None, columns=col_list)

for run in runs_for_three_days:
    df = pd.DataFrame(index=range(1,6),columns=col_list)
    df['date'][1] = run.date
    df['filename'][1] = run.logfile.name

    peaks1 = [sample.peak for sample in run.samples if sample.sample_num in [0,2,4,6,8]]
    peaks2 = [sample.peak for sample in run.samples if sample.sample_num in [1,3,5,7,9]]

    df.loc[0:5, 'peak1'] = [(peak.pa if peak else None) for peak in peaks1]
    df.loc[0:5, 'peak2'] = [(peak.pa if peak else None) for peak in peaks2]

    master_df = master_df.append(df)

append_df_to_excel(file, master_df, **{'index': False})


from openpyxl import load_workbook

wb = load_workbook(file)
ws = wb['Sheet1']

for num, row in enumerate(ws.iter_rows(min_row=ws.min_row+1, max_row=ws.max_row, min_col=5, max_col=5)):
    ws = set_formula_in_row(ws, num, row, mr_col=1)

for num, row in enumerate(ws.iter_rows(min_row=ws.min_row+1, max_row=ws.max_row, min_col=6, max_col=6)):
    ws = set_formula_in_row(ws, num, row, mr_col=2)

column_dimensions = {'A': 20,
                     'B': 22,
                     'C': 8,
                     'D': 8,
                     'E': 12,
                     'F': 12,
                     'G': 12,
                     'H': 8,
                     'I': 11,
                     'J': 8}

for col, val in column_dimensions.items():
    ws.column_dimensions[col].width = val

wb.save(file)