from pathlib import Path
import logging
from datetime import datetime
import datetime as dt
import statistics as s

from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column, Integer, String, Float, DateTime, ForeignKey
from sqlalchemy.orm import relationship

Base = declarative_base()  # needed to subclass for sqlalchemy objects

# retention times based on sample number
sample_rts = {0: [2, 3],  # TODO: Sample RTs should be treated similar to CompoundWindows in summit_voc
              1: [8.3, 9.3],  # even though methane RTs are quite stable, large GC changes will still need to be handled
              2: [14.65, 15.65],
              3: [21, 22],
              4: [27.3, 28.3],
              5: [33.65, 34.65],
              6: [40, 41],
              7: [46.4, 47.4],
              8: [52.75, 53.75],
              9: [59, 60]}


class Standard(Base):
    """
    A Standard is a container for the working standard on the methane GC. Standards have a start and end date, which
    specifies to which samples it should be applied. Ambient mixing ratios are calculated by the peak area ratio of
    a sample over the standard value, times the mixing ratio of the Standard that was in use during that period.

    Once a Standard is applied to a sample, it is related to it in a many-one relationship.
    """
    __tablename__ = 'standards'

    id = Column(Integer, primary_key=True)
    name = Column(String)
    mr = Column(Float)
    date_st = Column(DateTime)
    date_en = Column(DateTime)

    sample = relationship('Sample', back_populates='standard')

    def __init__(self, name, mr, date_st, date_en):
        self.name = name
        self.mr = mr
        self.date_st = date_st
        self.date_en = date_en


class Peak(Base):
    """
    A peak is just that, a signal peak in PeakSimple, Agilent, or another chromatography software. It is intiated in the
    creation of a PaLine, and gets related to the PaLine and any subsequent Samples or GcRuns that are matched to that
    PaLine.
    Peaks are created for ANY peak in a chromatogram, ie any rise above baseline above a low threshold for peak area.
    Peaks that are identified as actual sample/standard results are then assigned to a Sample.
    """

    __tablename__ = 'peaks'

    id = Column(Integer, primary_key=True)
    name = Column(String)
    pa = Column(Float)
    mr = Column(Float)
    rt = Column(Float)
    rev = Column(Integer)
    qc = Column(Integer)

    run = relationship('GcRun', back_populates='peaks')
    run_id = Column(Integer, ForeignKey('runs.id'))

    pa_line = relationship('PaLine', back_populates='peaks')
    pa_line_id = Column(Integer, ForeignKey('palines.id'))

    sample = relationship('Sample', uselist=False, back_populates='peak')

    def __init__(self, name, pa, rt):
        self.name = name.lower()
        self.pa = pa
        self.mr = None
        self.rt = rt
        self.rev = 0
        self.qc = 0

    def __str__(self):
        # Print the name, pa, and rt of a peak when called
        return f'<name: {self.name} pa: {self.pa} rt: {self.rt}, mr: {self.mr}>'

    def __repr__(self):
        # Print the name, pa, and rt of a peak when called
        return f'<name: {self.name} pa: {self.pa} rt: {self.rt}, mr: {self.mr}>'


class Sample(Base):
    """
    There are 10 Samples in every GcRun. Samples are created from log information that is output by the LabView VI for
    each run of the GC. Samples are later given a peak based on it's retention time and the order of the Samples.
    A Sample is not given a Peak if there is no Peak in the PaLine that sits in the retention time window.

    The quantifier of a Peak is a self-referencing relationship to the Sample that was used to calculate it's mixing
    ratio. For instance, Samples 1, 2, 4, and 5 in each run should be quantified by Sample 3 for that run, as the third
    sample is traditionally a standard injection. Samples 1, 2, 4, and 5 will therefore be related to Sample 3 in the
    'samples' table.
    """
    __tablename__ = 'samples'

    id = Column(Integer, primary_key=True)
    date = Column(DateTime, unique=True)

    peak = relationship('Peak', uselist=False, back_populates='sample')
    peak_id = Column(Integer, ForeignKey('peaks.id'))

    run = relationship('GcRun', back_populates='samples')
    run_id = Column(Integer, ForeignKey('runs.id'))

    standard = relationship('Standard', back_populates='sample')
    standard_id = Column(Integer, ForeignKey('standards.id'))

    quantifier = relationship('Sample', remote_side=[id])  # relate a sample to it's quantifying sample if applicable
    quantifier_id = Column(Integer, ForeignKey('samples.id'))

    correction = relationship('SampleCorrection', back_populates='sample')
    correction_id = Column(Integer, ForeignKey('corrections.id'))

    flow = Column(Float)
    pressure = Column(Float)
    rh = Column(Float)
    relax_p = Column(Float)
    sample_type = Column(Integer)
    sample_num = Column(Integer)

    def __init__(self, run, flow, pressure, rh, relax_p, sample_type, sample_num):
        self.run = run
        self.flow = flow
        self.pressure = pressure
        self.rh = rh
        self.relax_p = relax_p
        self.sample_type = sample_type
        self.sample_num = sample_num


class SampleCorrection(Base):

    __tablename__ = 'corrections'

    id = Column(Integer, primary_key=True)
    pa = Column(Float)
    sample_num = Column(Integer)

    sample = relationship('Sample', uselist=False, back_populates='correction')

    def __init__(self, sample_num, pa, sample):
        self.sample_num = sample_num
        self.pa = pa
        self.sample = sample


class PaLine(Base):
    """
    A PaLine is a line in CH4.LOG that contains some number of peaks that were integrated by PeakSimple. A PaLine and
    multiple Samples compose a GcRun.
    """
    __tablename__ = 'palines'

    id = Column(Integer, primary_key=True)
    peaks = relationship('Peak', back_populates='pa_line')
    run = relationship('GcRun', back_populates='pa_line')
    date = Column(DateTime, unique=True)
    status = Column(String)

    def __init__(self, date, peaks):
        self.date = date
        self.peaks = peaks
        self.status = 'single'


class GcRun(Base):
    """
    A GcRun is created from reading the LabView VI files at the same time as the Samples it contains. GcRuns will
    contain 10 Samples assuming the VI is not changed, though each Sample may not have a Peak associated if the
    injection failed or a peak could not be integrated.

    Once all the Samples in a GcRun have been given Peaks and quantified, a GcRun will be given a median, and stdev,
    based on the quantified peaks it's related to.
    """

    __tablename__ = 'runs'

    id = Column(Integer, primary_key=True)
    peaks = relationship('Peak', back_populates='run')
    samples = relationship('Sample', back_populates='run')
    pa_line = relationship('PaLine', back_populates='run')
    pa_line_id = Column(Integer, ForeignKey('palines.id'))
    median = Column(Float)
    rsd = Column(Float)  # relative standard deviation (stdev/median) * 100
    standard_rsd = Column(Float)

    _logfile = Column(String)
    date = Column(DateTime, unique=True)
    carrier_flow = Column(Float)
    sample_flow = Column(Float)
    sample_time = Column(Float)
    relax_time = Column(Float)
    injection_time = Column(Float)
    wait_time = Column(Float)
    loop_p_check1 = Column(Float)
    loop_p_check2 = Column(Float)
    status = Column(String)

    def __init__(self, logfile, date, carrier_flow, sample_flow, sample_time, relax_time,
                 injection_time, wait_time, loop_p_check1, loop_p_check2):
        self.carrier_flow = carrier_flow
        self.sample_flow = sample_flow
        self.sample_time = sample_time
        self.relax_time = relax_time
        self.injection_time = injection_time
        self.wait_time = wait_time
        self.loop_p_check1 = loop_p_check1
        self.loop_p_check2 = loop_p_check2
        self.logfile = logfile
        self.date = date
        self.status = 'single'

    @property  # lets logfile be stored as String, but returns Path object
    def logfile(self):
        return Path(self._logfile)

    @logfile.setter
    def logfile(self, value):
        self._logfile = str(value)


class Datum():
    """
    A QC'd version of a peak...? Maybe, not sure yet.
    """
    pass


def read_pa_line(line):
    """
    Takes one line as a string from a PeakSimple log, and processes it in Peak objects and an NmhcLine containing those
    peaks. (This one is a minor modification of the one in summit_voc.py)
    :param line: string, one line of data from VOC.LOG, NMHC_PA.LOG, etc.
    :return: NmhcLine or None
    """

    ls = line.split('\t')
    line_peaks = []

    line_date = datetime.strptime(ls[1] + ' ' + ls[2], '%m/%d/%Y %H:%M:%S')

    for ind, item in enumerate(ls[3:]):

        ind = ind + 3  # offset ind since we're working with ls[3:]

        peak_dict = dict()

        if '"' in item:

            peak_dict['name'] = item.strip('"')  # can't fail, " is definitely there

            try:
                peak_dict['rt'] = float(ls[ind + 1])
            except:
                peak_dict['rt'] = None
            try:
                peak_dict['pa'] = float(ls[ind + 2])
            except:
                peak_dict['pa'] = None

            if None not in peak_dict.values():
                line_peaks.append(Peak(peak_dict['name'], peak_dict['pa'], peak_dict['rt']))

    if len(line_peaks) == 0:
        this_line = None
    else:
        this_line = PaLine(line_date, line_peaks)

    return this_line


def read_log_file(path):
    """
    A single log file actually accounts for a whole run, which includes 10 samples. There is information that is
    run-only, then sample-specific data. These are parsed into a GcRun containing 10 Samples.

    :param path: pathlib Path to file
    :return: GcRun, containing 10 samples and metadata
    """

    contents = path.read_text().split('\n')

    # filter out blank line the VI is writing at the end of files that read_text() keeps
    contents[:] = [line for line in contents if line != '']

    run_year = int(path.name[:4])
    run_doy = int(path.name[4:7]) - 1  # when adding, if it's DOY 1, you don't want to add 1 to Jan 1 of that year...
    run_hour = int(path.name[7:9])

    run_date = datetime(run_year, 1, 1) + dt.timedelta(days=run_doy, hours=run_hour)

    run_dict = {}
    run_dict['logfile'] = path
    run_dict['date'] = run_date
    run_dict['carrier_flow'] = float(contents[0].split('\t')[1])
    run_dict['sample_flow'] = float(contents[1].split('\t')[1])
    run_dict['sample_time'] = float(contents[2].split('\t')[1])
    run_dict['relax_time'] = float(contents[3].split('\t')[1])
    run_dict['injection_time'] = float(contents[4].split('\t')[1])
    run_dict['wait_time'] = float(contents[5].split('\t')[1])
    run_dict['loop_p_check1'] = float(contents[-2].split('\t')[1])
    run_dict['loop_p_check2'] = float(contents[-1].split('\t')[1])

    run = GcRun(**run_dict)

    # run information is contained in 23-line blocks with no delimiters, spanning (0-indexed) lines 17:-3
    # each block of 23 will
    run_blocks = contents[17:-2]
    # run blocks come in sets of 23 lines with no particular delimiters
    indices = [i * 23 for i in range(int(len(run_blocks) / 23))]

    samples = []
    for num, ind in enumerate(indices):
        sample_info = run_blocks[ind:ind + 23]
        sample_dict = {}
        sample_dict['flow'] = float(sample_info[0].split('\t')[1])
        sample_dict['pressure'] = float(sample_info[1].split('\t')[1])
        sample_dict['rh'] = float(sample_info[2].split('\t')[1])
        sample_dict['relax_p'] = s.mean([float(sample_info[i].split('\t')[1]) for i in range(3, 23)])
        sample_dict['sample_type'] = int(float(contents[6 + num].split('\t')[1]))  # get sample type (int) from file
        sample_dict['sample_num'] = num  # assign number of sample in the GC sequence

        samples.append(Sample(run, **sample_dict))

    run.samples = samples

    return run


def match_lines_to_runs(lines, runs):
    """
    This takes a list of PaLine and GcRun objects and matched them by date, within a tolerance.
    When matching objects, it WILL modify their parameters and status if warranted.

    :param lines: list, of PaLine objects that are unmatched
    :param runs: list, of GcRun objects that are unmatched
    :return: (lines, runs, match_count) list of line objects, list of run objects, and int of runs that were matched
    """
    match_count = 0
    from summit_core import find_closest_date, search_for_attr_value

    logger = logging.getLogger(__name__)

    for line in lines:
        # For each log, attempt to find matching NmhcLine
        # unpack date attr from all NmhcLines provided
        run_dates = [run.date for run in runs]

        [match, diff] = find_closest_date(line.date, run_dates)  # get matching date and it's difference

        if not match or not diff:
            continue

        if abs(diff) < dt.timedelta(minutes=70):
            # Valid matches *usually* *HAD* ~03:22 difference
            # on 6/12/2019, the sequence was changed, which resulted in differences ranging from 15 min to 55 min.
            # on 6/20/2019 a handful of runs were ~60 min after, so the tolerance was upped to 70 min

            matched_run = search_for_attr_value(runs, 'date', match)  # pull matching NmhcLine

            line.status = 'married'
            matched_run.status = 'married'

            for peak in line.peaks:
                peak.run = matched_run  # relate all peaks in pa line to the newly matched run

            matched_run.pa_line = line
            logger.info(f'PaLine {line.date} matched to GcRun for {matched_run.date}.')
            match_count += 1
        else:
            continue

    return (lines, runs, match_count)


def calc_ch4_mr(sample, quantifier, standard):
    """
    Calculate mixing ratio of a sample when provided with the sample to quantify it against, and the Standard that was
    run in the quantifier.

    :param sample: Sample, a valid sample (use valid_sample(Sample) to verify)
    :param quantifier: Sample, a valid sample ("") that will be used to quantify the provided sample
    :param standard: Standard, the standard that was run.
    :return: Sample, the same object but returned with it's peak given a mixing ratio
    """
    sample.quantifier = quantifier
    sample.peak.mr = (sample.peak.pa / sample.quantifier.peak.pa) * standard.mr
    sample.standard = standard

    return sample


def valid_sample(sample):
    """
    Check if sample will allow quantification. First check not None, then it has a peak and the peak is not None.
    :param sample: Sample, to check for validity
    :return: Bool, True if the sample is valid
    """
    if sample is None or sample.peak is None or sample.peak.pa is None:
        return False
    else:
        return True


def plottable_sample(sample):
    """

    TODO - Rework to include tighter controls

    Check sample for validity, then check for a mixing ratio between reasonable bounds.

    :param sample: Sample, to check if it should appear in plots
    :return: Boolean, True if sample should be plotted
    """
    if not valid_sample(sample):
        return False
    else:
        if sample.peak.mr is None or not (1850 < sample.peak.mr < 2050):
            return False
        else:
            return True


def set_formula_in_row(ws, num, row, mr_col=1):
    """
    This "loops" through single cells in a column and applies mixing ratio formulas. The format of the formulas is quite
    convulted for legacy reasons. Existing procedures made adhering to this format easier, but the gist is below.

    Samples come in sets of 10, from a GC Run. They were kept in two columns in a spreadsheet, where the below is
    repeated in sets of 5 rows (per GC Run). One GC Run looks like:

    col1  | col2
    -------------
    samp1 | samp2
    std1  | samp4
    samp5 _______  # all beyond this line are quantified with standard 2
    ______| samp6
    samp7 | std2
    samp9 | samp10

    Samples 1-5 use the first standard (sample 3) to quantify themselves, and samples 6-10 use the second standard
    (sample 8) to quantify themselves. These rules are applied in the integration code itself, but the mixing ratios
    and relevant statistics need to be calculated within the spreadsheet so the person integrating has access to them as
    they integrate manually.

    The sheet is loaded by add_formulas_and_format_sheet(), then the two columns are passed row-by-row to this function
    to add the formulae before saving.

    :param ws: object, open worksheet with openpyxl as the engine
    :param num: int, absolute row number (excluding header
    :param row: object, the row object generated by iterating over ws
    :param mr_col: int, in [1,2]
    :return: ws, the modified worksheet is passed back
    """

    assert mr_col in [1,2], "Invalid mixing ratio column. It must either 1 or 2"

    std_relnum = 1 if mr_col is 1 else 3
    # if it's the first mixing ratio column, the standard will be in the second row (0-indexed: 1)
    # if it's the second mixing ratio columnm, the standard will be in the fourth row (0-indexed: 3)

    standard_div_line = 2 if mr_col is 1 else 1
    # samples 1-5 (excluding the standardd) are quantified using the first standard (sample 3)
    # samples 6-10 (excluding the stnadrad) are quantified using the second standard (sample 8)
    # so, in column 1, every sample up to (0-indexed) 2 should be quantified with standard 1, and
    # everything after is quantified with standard 2. In column 2, that number changes to 1

    relnum = num % 5
    # num is 0-indexed, relnum is the position in this group of 5 rows (one run is 5r x 2c for 10 total runs)
    if relnum is std_relnum: return ws  # skip the standard for this column

    for cell in row:
        if cell.value is None:  # assume cells with some value have been modified and should not be changed
            rownum = cell.row  # retrieve the real row number
            pa_cell = f'C{rownum}' if mr_col is 1 else f'D{rownum}'
            # the peak area for a mixing ratio cell will always be C for column 1 and D for column 2, always same row

            if relnum <= standard_div_line:  # this is should be quantified by standard 1, in this column
                std_pa_cell = f'C{rownum - relnum + 1}'
            else:  # it should be quantified by standard 2, in the next column
                std_pa_cell = f'D{rownum - relnum + 3}'

            cell.value = f'={pa_cell}/{std_pa_cell} * 2067.16'

            if relnum is 0 and mr_col is 1:  # the first line of every 5-row batch needs additional statistics added
                # this does not need to be done twice, which is why it's done only for MR col 1

                run_range = f'E{rownum}:F{rownum+4}'  # all mixing cells in the run
                std_range = f'C{rownum+1}, D{rownum+3}'  # the two standards

                run_median_cell = ws[f'G{rownum}']
                run_rsd_cell = ws[f'H{rownum}']
                std_med_cell = ws[f'I{rownum}']
                std_rsd_cell = ws[f'J{rownum}']

                run_rsd_cell.number_format = '0.00%'
                std_rsd_cell.number_format = '0.00%'

                run_median_cell.value = f'=MEDIAN({run_range})'  # set formulas
                run_rsd_cell.value = f'=STDEV({run_range})/{run_median_cell.coordinate}'
                std_med_cell.value = f'=MEDIAN({std_range})'
                std_rsd_cell.value = f'=STDEV({std_range})/{std_med_cell.coordinate}'
    return ws


def add_formulas_and_format_sheet(filename):
    """
    Loads the methane spreadsheet created by a process and adds formulas if they're not present.
    Column widths are also adjusted for readability.
    :param filename: Path-like object of the file to be read
    :return: None
    """
    from openpyxl import load_workbook
    wb = load_workbook(filename)
    ws = wb['Sheet1']

    for num, row in enumerate(ws.iter_rows(min_row=ws.min_row + 1, max_row=ws.max_row, min_col=5, max_col=5)):
        ws = set_formula_in_row(ws, num, row, mr_col=1)

    for num, row in enumerate(ws.iter_rows(min_row=ws.min_row + 1, max_row=ws.max_row, min_col=6, max_col=6)):
        ws = set_formula_in_row(ws, num, row, mr_col=2)

    column_dimensions = {'A': 20,
                         'B': 22,
                         'C': 8,
                         'D': 8,
                         'E': 12,
                         'F': 12,
                         'G': 12,
                         'H': 8,
                         'I': 11,
                         'J': 8}

    for col, val in column_dimensions.items():
        ws.column_dimensions[col].width = val

    wb.save(filename)
    return


def summit_methane_plot(dates, compound_dict, title = None, limits=None,
                        minor_ticks=None, major_ticks=None, unit_string='ppbv'):
    """
    :param dates: list, of Python datetimes; if set, this applies to all compounds.
        If None, each compound supplies its own date values
    :param compound_dict: dict, {'compound_name':[dates, mrs]}
        keys: str, the name to be plotted and put into filename
        values: list, len(list) == 2, two parallel lists that are...
            dates: list, of Python datetimes. If None, dates come from dates input parameter (for all compounds)
            mrs: list, of [int/float/None]s; these are the mixing ratios to be plotted
    :param limits: dict, optional dictionary of limits including ['top','bottom','right','left']
    :param minor_ticks: list, of major tick marks
    :param major_ticks: list, of minor tick marks
    :param unit_string: string, will be displayed in y-axis label as f'Mixing Ratio ({unit_string})'
    :return: None

    This plots stuff.

    Example with all dates supplied:
        plot_last_week((None, {'Ethane':[[date, date, date], [1, 2, 3]],
                                'Propane':[[date, date, date], [.5, 1, 1.5]]}))

    Example with single date list supplied:
        plot_last_week([date, date, date], {'ethane':[None, [1, 2, 3]],
                                'propane':[None, [.5, 1, 1.5]]})
    """

    import matplotlib.pyplot as plt
    from matplotlib.dates import DateFormatter
    from pandas.plotting import register_matplotlib_converters
    register_matplotlib_converters()

    f1 = plt.figure()
    ax = f1.gca()

    if dates is None:  # dates supplied by individual compounds
        for compound, val_list in compound_dict.items():
            assert val_list[0] is not None, 'A supplied date list was None'
            assert ((len(val_list[0]) > 0) and (len(val_list[0]) == len(val_list[1]))), \
                'Supplied dates were empty or lengths did not match'

            ax.plot(val_list[0], val_list[1], '-o')

    else:
        for compound, val_list in compound_dict.items():
            ax.plot(dates, val_list[1], '-o')

    compounds_safe = []
    for k, _ in compound_dict.items():
        """Create a filename-safe list using the given legend items"""
        compounds_safe.append(k.replace('-', '_').replace('/', '_').lower())

    comp_list = ', '.join(compound_dict.keys())  # use real names for plot title
    fn_list = '_'.join(compounds_safe)  # use 'safe' names for filename

    if limits is not None:
        ax.set_xlim(right=limits.get('right'))
        ax.set_xlim(left=limits.get('left'))
        ax.set_ylim(top=limits.get('top'))
        ax.set_ylim(bottom=limits.get('bottom'))

    if major_ticks is not None:
        ax.set_xticks(major_ticks, minor=False)
    if minor_ticks is not None:
        ax.set_xticks(minor_ticks, minor=True)

    date_form = DateFormatter("%Y-%m-%d")
    ax.xaxis.set_major_formatter(date_form)

    [i.set_linewidth(2) for i in ax.spines.values()]
    ax.tick_params(axis='x', labelrotation=30)
    ax.tick_params(axis='both', which='major', size=8, width=2, labelsize=15)
    f1.set_size_inches(11.11, 7.406)

    if not title:
        title = f'{comp_list}'

    ax.set_ylabel(f'Mixing Ratio ({unit_string})', fontsize=20)
    ax.set_title(title, fontsize=24, y=1.02)
    ax.legend(compound_dict.keys())

    f1.subplots_adjust(bottom=.20)

    plot_name = f'{fn_list}_last_week.png'.replace(' ', '_')

    f1.savefig(plot_name, dpi=150)
    plt.close(f1)

    return plot_name
