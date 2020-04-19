import os
import json
import asyncio
from pathlib import Path
import datetime as dt
from datetime import datetime

from sqlalchemy.types import TypeDecorator, VARCHAR
from sqlalchemy.ext.mutable import MutableDict, MutableList
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column, Integer, String, Boolean, DateTime

Base = declarative_base()

PROC = 'Core'


def find_project_dir(runpath):
    """
    Scans up directories until it finds the project folder.
    :param runpath: pathlib.Path, where summit_core is called from.
    :return: pathlib.Path, the base project directory
    """
    runpath = runpath.resolve()

    if runpath.name == "Summit" or runpath.name == 'summit_master':
        return runpath
    else:
        runpath = runpath / '..'
        return find_project_dir(runpath)


project_dir = find_project_dir(Path(os.getcwd()))

voc_dir = project_dir / 'processors/summit_voc_processor'
picarro_dir = project_dir / 'processors/summit_picarro_processor'
methane_dir = project_dir / 'processors/summit_methane_processor'
error_dir = project_dir / 'processors/errors'
core_dir = project_dir / 'core'
taylor_basepath = '/data/web/htdocs/instaar/groups/arl/res_parameters/summit_plots'

processor_dirs = [voc_dir, picarro_dir, methane_dir, error_dir, core_dir]

data_file_paths = json.loads((core_dir / 'file_locations.json').read_text())

for k, v in data_file_paths.items():
    data_file_paths[k] = Path(v)  # Pathify stored string paths

methane_LOG_path = data_file_paths.get('methane_LOG')
voc_LOG_path = data_file_paths.get('voc_LOG')

methane_logs_path = data_file_paths.get('methane_logs')
voc_logs_path = data_file_paths.get('voc_logs')
daily_logs_path = data_file_paths.get('daily_logs')
picarro_logs_path = data_file_paths.get('picarro_logs')

methane_logs_sync = data_file_paths.get('methane_logs_sync')
voc_logs_sync = data_file_paths.get('voc_logs_sync')
daily_logs_sync = data_file_paths.get('daily_logs_sync')
picarro_logs_sync = data_file_paths.get('picarro_logs_sync')

taylor_auth = data_file_paths.get('taylor_server_auth')


class Config(Base):
    """
    Configs are a somewhat generic storage container for configuration information for each processor. Configs are kept
    in the core database, and not every column is used by every processor. Some are used loosely, while others were
    designed specifically for the processor they're used exclusively in. Default values are given so only the necessary
    new values need to be given for init.
    """
    __tablename__ = 'config'

    id = Column(Integer, primary_key=True)

    processor = Column(String, unique=True)  # only one config per processor
    filesize = Column(Integer)
    pa_startline = Column(Integer)
    last_data_date = Column(DateTime)
    days_to_plot = Column(Integer)

    def __init__(self, processor=None, filesize=0, pa_startline=0, last_data_date=datetime(1900, 1, 1), days_to_plot=7):
        self.processor = processor
        self.filesize = filesize
        self.pa_startline = pa_startline
        self.last_data_date = last_data_date
        self.days_to_plot = days_to_plot


class TempDir:
    """
    Context manager for working in a directory.
    """

    def __init__(self, path):
        self.old_dir = os.getcwd()
        self.new_dir = path

    def __enter__(self):
        os.chdir(self.new_dir)

    def __exit__(self, *args):
        os.chdir(self.old_dir)


class JDict(TypeDecorator):
    """
    Serializes a dictionary for SQLAlchemy storage.
    """
    impl = VARCHAR

    def process_bind_param(self, value, dialect):
        if value is not None:
            value = json.dumps(value)
        return value

    def process_result_value(self, value, dialect):
        if value is not None:
            value = json.loads(value)
        return value


class JList(TypeDecorator):
    """
    Serializes a list for SQLAlchemy storage.
    """
    impl = VARCHAR

    def process_bind_param(self, value, dialect):
        value = json.dumps(value)
        return value

    def process_result_value(self, value, dialect):
        value = json.loads(value)
        return value


class Plot(Base):
    """
    Plots are used to register created plots as ready to upload in the core database. When run, individual processors
    will log Plots to the core database with their filepath, remote path they should be sent to, and staged == True if
    they should be uploaded. check_send_plots() then loads these from the core database all at once and sends are staged
    plots to their specified remote paths.
    """
    __tablename__ = 'plots'

    id = Column(Integer, primary_key=True)

    staged = Column(Boolean)
    _path = Column(String, unique=True)
    remote_path = Column(String)
    _name = Column(String)

    def __init__(self, path, remote_path, staged):
        self.path = path
        self.remote_path = remote_path
        self.staged = staged

    @property
    def path(self):
        return Path(self._path)

    @path.setter
    def path(self, value):
        self._path = str(value.resolve())
        self._name = value.name

    @property
    def name(self):
        return self._name


MutableList.associate_with(JList)
MutableDict.associate_with(JDict)


def configure_logger(rundir, name):
    """
    Create the project-specific logger. DEBUG and up is saved to the log, INFO and up appears in the console.

    :param rundir: Path, to create log sub-path in
    :param name: str, name for logfile
    :return: logger object
    """
    from pathlib import Path
    import logging

    logfile = Path(rundir) / f'processor_logs/{name}.log'
    logger = logging.getLogger(name)
    logger.setLevel(logging.DEBUG)
    fh = logging.FileHandler(logfile)
    fh.setLevel(logging.DEBUG)

    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)

    formatter = logging.Formatter('%(asctime)s -%(levelname)s- %(message)s')

    [H.setFormatter(formatter) for H in [ch, fh]]
    if not len(logger.handlers):
        _ = [logger.addHandler(H) for H in [ch, fh]]

    return logger


def connect_to_db(engine_str, directory):
    """
    Takes string name of the database to create/connect to, and the directory it should be in.

    :param engine_str: connection string for the database
    :param directory: directory the database should in (created?) in
    :return: engine, session, Base

    Example:
    engine, session, Base = connect_to_db('sqlite:///reservoir.sqlite', dir)
    """

    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    # Base = declarative_base()  # needed to subclass for sqlalchemy objects

    with TempDir(directory):
        engine = create_engine(engine_str)
    sessy = sessionmaker(bind=engine)
    sess = sessy()

    return engine, sess


def check_path_date(filepath):
    """
    Returns the file's 'Date Modified' from window
    :param filepath: file-like object
    :return: datetime object of date modified
    """

    if Path.is_file(filepath):
        epochseconds = Path.stat(filepath).st_mtime
        return dt.datetime.fromtimestamp(epochseconds)


def check_filesize(filepath):
    """
    Returns the filesize in bytes.
    :param filepath: file-like object
    :return: int, filesize in bytes
    """
    import logging

    logger = logging.getLogger(__name__)

    if Path.is_file(filepath):
        return Path.stat(filepath).st_size
    else:
        logger.warning(f'File {filepath.name} not found.')
        return


def list_files_recur(path):
    """
    :param path: pathlib Path object
    :return: list, of file-like Path objects
    """
    files = []
    for file in path.rglob('*'):
        files.append(file)

    return files


def get_all_data_files(path, filetype):
    """
    Recursively search the given directory for .xxx files.

    :param path: Path to search
    :param filetype: str, ".type" of file to search for
    :return: list, of file-like Path objects
    """
    files = list_files_recur(path)
    files[:] = [file for file in files if filetype in file.name]

    return files


def merge_lists(a, b):
    """
    Generator to merge the lists a, and b, starting with the first element in a.
    :param a: list, or other iterable
    :param b: list, or other iterable
    :return: list, joined from a and b, such that the return is [a[0], b[0], a[1], b[1], ...]
    """
    assert len(a) == len(b), "Lists of non-matching lengths cannot be merged."
    it_a = iter(a)
    it_b = iter(b)
    for a, b in zip(it_a, it_b):
        yield a
        yield b


def split_into_sets_of_n(lst, n):
    """
    Split a list into lists of length n
    :param lst: list, a list of values to be split
    :param n: int, number of items per set after split
    :return: generator, lists of length n
    """
    for i in range(0, len(lst), n):
        yield (lst[i:i + n])


def search_for_attr_value(obj_list, attr, value):
    """
    Finds the first (not necesarilly the only) object in a list, where its
    attribute 'attr' is equal to 'value', returns None if none is found.
    :param obj_list: list, of objects to search
    :param attr: string, attribute to search for
    :param value: mixed types, value that should be searched for
    :return: obj, from obj_list, where attribute attr matches value
        **** warning: returns the *first* obj, not necessarily the only
    """
    return next((obj for obj in obj_list if getattr(obj, attr, None) == value), None)


def find_closest_date(date, list_of_dates, how='abs'):
    """
    This is a helper function that works on Python datetimes. It returns the closest date value, either absolutely,
    or closest without being above/below, and the timedelta from the provided date.

    :param date: datetime
    :param list_of_dates: list, of datetimes
    :param how: ['abs', 'pos','neg']
            'abs': Absolute closest datetime, either above or below the given date
            'pos': Matched datetime must be greater than given date
            'neg': Matched datetime must be less than given date
    :return: match, delta: the matching date from the list, and it's difference to the original as a timedelta
    """

    if how == 'abs':
        try:
            match = min(list_of_dates, key=lambda x: abs(x - date))
        except ValueError:
            return None, None
    elif how == 'pos':
        try:
            deltas = [list_date - date for list_date in list_of_dates if list_date - date >= dt.timedelta(0)]
            match = min(deltas)
        except ValueError:
            return None, None
    elif how == 'neg':
        try:
            deltas = [list_date - date for list_date in list_of_dates if list_date - date <= dt.timedelta(0)]
            match = max(deltas)
        except ValueError:
            return None, None
    else:
        assert False, "Supplied 'how' not in ['abs', 'pos', 'neg']"

    delta = match - date

    return match, delta


def create_daily_ticks(days_in_plot, minors_per_day=4):
    """
    Takes a number of days to plot back, and creates major (1 day) and minor (6 hour) ticks.

    :param days_in_plot: int, number of days to be displayed on the plot
    :param minors_per_day: int, number of minor ticks per day
    :return: date_limits, major_ticks, minor_ticks
    """
    from datetime import datetime

    date_limits = dict()
    date_limits['right'] = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + dt.timedelta(
        days=1)  # end of day
    date_limits['left'] = date_limits['right'] - dt.timedelta(days=days_in_plot)

    major_ticks = [date_limits['right'] - dt.timedelta(days=x) for x in range(0, days_in_plot + 1)]

    minor_ticks = [date_limits['right'] - dt.timedelta(hours=x * (24 / minors_per_day))
                   for x in range(0, days_in_plot * minors_per_day + 1)]

    return date_limits, major_ticks, minor_ticks


def five_minute_medians(dates, vals):
    """
    Takes a list of dates and matching list of values, averaging them to 5-minute intervals and returns.
    :param dates: list, of datetimes
    :param vals: list, of any type that can be put in a dataframe
    :return: list, list
    """
    import pandas as pd
    df = pd.DataFrame({'dates': dates, 'vals': vals})
    df.set_index(df['dates'], inplace=True, drop=True)
    df = df.resample('5min').median()

    return df.index.tolist(), df['vals'].tolist()


def connect_to_sftp():
    """
    Uses paramiko to create a connection to the Taylor drive. Relies on authetication information from a JSON file.
    :return: SFTP_Client
    """
    import paramiko
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    server_info = json.loads(taylor_auth.read_text())
    client.connect(**server_info)
    return client.open_sftp()


def add_or_ignore_plot(plot, core_session):
    """
    Queries database and adds a plot only if one with the same path doesn't already exist.
    :param plot: Plot, to be added to database
    :param core_session: sqlalchemy.Session object
    :return: None
    """
    plots_in_db = [p[0] for p in core_session.query(Plot._path).all()]

    if str(plot.path.resolve()) not in plots_in_db:
        core_session.add(plot)
    return


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename] into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Adapted From:
        https://stackoverflow.com/questions/38074678/append-existing-excel-sheet-with-new-dataframe-using-python-pandas
    :param filename: File path or existing ExcelWriter (Example: '/path/to/file.xlsx')
    :param df: dataframe to save to workbook
    :param sheet_name: Name of sheet which will contain DataFrame. (default: 'Sheet1')
    :param startrow: upper left cell row to dump data frame. Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
    :param truncate_sheet: truncate (remove and recreate) [sheet_name] before writing DataFrame to Excel file
    :param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()` [can be dictionary]
    :return: None
    """
    import pandas as pd
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0
    else:
        # If there's a start row, don't write the header when appending
        if not to_excel_kwargs:
            to_excel_kwargs = {'header': False}
        else:
            to_excel_kwargs['header'] = False

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    return


async def send_files_sftp(filepaths, remote_path):
    """
    Send a list of files to the provided remote path.
    :param filepaths: list, of pathlib Path objects
    :param remote_path: string, path on the remote server to send files to
    :return: list, of booleans of which plots uploaded sucessfully
    """

    bools = []
    try:
        con = connect_to_sftp()
        con.chdir(remote_path)

        for file in filepaths:
            try:
                con.put(str(file), file.name)
                bools.append(True)
            except:
                bools.append(False)

        con.close()
    except:
        while len(bools) < len(filepaths):
            bools.append(False)

    return bools


async def check_send_plots(logger):
    """
    Look through all plots staged to be uploaded and remove them if successfully uploaded.
    :param logger: logging logger to log to
    :return: boolean, True if ran without errors
    """
    try:
        from summit_errors import send_processor_email
    except ImportError:
        logger.error('ImportError occurred in check_send_plots()')
        return False

    try:
        engine, session = connect_to_db('sqlite:///summit_core.sqlite', core_dir)
    except Exception as e:
        logger.error(f'Exception {e.args} prevented connection to the database in check_send_plots()')
        send_processor_email('Core', exception=e)
        return False

    try:
        plots_to_upload = session.query(Plot).filter(Plot.staged == True)

        remote_dirs = set([p.remote_path for p in plots_to_upload.all()])

        for remote_dir in remote_dirs:
            plot_set = plots_to_upload.filter(Plot.remote_path == remote_dir).all()

            if plot_set:
                paths_to_upload = [p.path for p in plot_set]
                successes = await send_files_sftp(paths_to_upload, remote_dir)

                for plot, success in zip(plots_to_upload, successes):
                    if success:
                        logger.info(f'Plot {plot.name} uploaded to website.')
                        session.delete(plot)
                    else:
                        logger.warning(f'Plot {plot.name} failed to upload.')

        session.commit()

        session.close()
        engine.dispose()
        return True

    except Exception as e:
        logger.error(f'Exception {e.args} occurred in check_send_plots().')
        send_processor_email('Core', exception=e)
        session.close()
        engine.dispose()
        return False


class MovedFile(Base):
    """
    MovedFiles are used to track files that have been moved to the /data directory from their FTP directories.
    The FTP directories are cleaned somewhat regularly, so they're not a good home for the permanent data files.

    Filepaths are used to track and move them, and filesize is the sole determiner if a file is moved again.
    """
    __tablename__ = 'files'

    id = Column(Integer, primary_key=True)

    _path = Column(String, unique=True)
    _name = Column(String)
    location = Column(String)
    size = Column(Integer)
    type = Column(String)

    def __init__(self, path, type, location, size):
        self.path = path
        self.type = type
        self.location = location  # either 'sync' or 'data'
        self.size = size

    @property
    def path(self):
        return Path(self._path)

    @path.setter
    def path(self, value):
        self._path = str(value)
        self._name = value.name

    @property
    def name(self):
        return self._name


async def move_log_files(logger):
    """
    Runs continuously and sleeps for 10 minutes at a time. Comb the directories for new data files and move any that
    are new or have been updated. This WILL NOT handle turning over a new year in the daily files well, as they have no
    year in the filename. I can't fix that.

    :param logger: logging logger to log to
    :return: boolean, True if ran without errors
    """

    while True:
        try:
            from summit_errors import send_processor_email, EmailTemplate, sender, processor_email_list
            from shutil import copy
            import datetime as dt
            import os
        except ImportError:
            logger.error('ImportError occurred in move_log_files()')
            return False

        try:
            engine, session = connect_to_db('sqlite:///summit_core.sqlite', core_dir)
            MovedFile.__table__.create(engine, checkfirst=True)
        except Exception as e:
            logger.error(f'Exception {e.args} prevented connection to the database in move_log_files()')
            send_processor_email('Core', exception=e)
            return False

        try:
            logger.info('Running move_log_files()')

            sync_paths = [methane_logs_sync, voc_logs_sync, daily_logs_sync, picarro_logs_sync]
            data_paths = [methane_logs_path, voc_logs_path, daily_logs_path, picarro_logs_path]
            data_types = ['methane', 'voc', 'daily', 'picarro']
            file_types = ['.txt', '.txt', '.txt', '.dat']

            for sync_path, type_, data_path, file_type in zip(sync_paths, data_types, data_paths, file_types):

                # change the name of the daily files before reading them in (implemented: 2/14/2020)
                for d in get_all_data_files(daily_logs_sync, '.txt'):
                    if check_path_date(d).year == dt.datetime.now().year and "2020" not in str(d):
                        name, extension = os.path.splitext(d)
                        d.rename(name + '_' + str(dt.datetime.now().year) + extension)

                sync_files = [MovedFile(path, type_, 'sync', check_filesize(path))
                              for path in get_all_data_files(sync_path, file_type)]

                data_files = (session.query(MovedFile)
                              .filter(MovedFile.location == 'data')
                              .filter(MovedFile.type == type_)
                              .all())
                moved_data_files = [d.name for d in data_files]

                for file in sync_files:
                    if file.name not in moved_data_files:
                        try:
                            copy(file.path, data_path)  # will overwrite
                        except PermissionError:
                            logger.error(f'File {file.name} could not be moved due to a permissions error.')
                            from summit_errors import send_processor_warning
                            send_processor_warning(PROC, 'PermissionError',
                                                   f'File {file.name} could not be moved due a permissions error.\n'
                                                   + 'Copying/pasting the file, deleting the old one, and renaming '
                                                   + 'the file to its old name should allow it to be processed.\n'
                                                   + 'This will require admin privelidges.')
                            continue
                        file.path = data_path / file.name
                        file.location = 'data'
                        session.merge(file)
                        logger.info(f'File {file.name} moved to data directory.')
                    else:
                        matched_file = search_for_attr_value(data_files, 'name', file.name)
                        if file.size > matched_file.size:
                            try:
                                copy(file.path, data_path)  # will overwrite
                            except PermissionError:
                                logger.error(f'File {file.name} could not be moved due to a permissions error.')
                                from summit_errors import send_processor_warning

                                send_processor_warning(PROC, 'PermissionError',
                                                       f'File {file.name} could not be moved due a permissions error.\n'
                                                       + 'Copying/pasting the file, deleting the old one, and renaming '
                                                       + 'the file to its old name should allow it to be processed.\n'
                                                       + 'This will require admin privelidges.')
                                continue

                            matched_file.size = check_filesize(matched_file.path)
                            session.merge(matched_file)
                            logger.info(f'File {matched_file.name} updated in data directory.')

            session.commit()

            session.close()
            engine.dispose()

            import gc
            gc.collect()

            for i in range(20):
                await asyncio.sleep(30)

        except Exception as e:
            logger.error(f'Exception {e.args} occurred in move_log_files().')
            send_processor_email('Core', exception=e)
            session.close()
            engine.dispose()
            return False
